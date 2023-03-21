import openpyxl
import tkinter 
from tkinter import Tk
import tkinter as tk
from tkinter import ttk, Checkbutton, LabelFrame, RIDGE, W, BooleanVar

def main():
    root = Tk()
    root.state('zoomed') 
    def todo():
        # Open the workbook
        workbook = openpyxl.load_workbook('Plany P15 - v.2 .xlsm', keep_vba=True)

        # Get the active worksheet
        worksheet = workbook['290 Act. 50']

        to_do_lf = LabelFrame(root, text='To-Do', bd=2, relief=RIDGE)

        
        # Create a canvas with a scrollbar
        canvas = tkinter.Canvas(to_do_lf, width=800, height=350)
        scrollbar = ttk.Scrollbar(to_do_lf, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")
            )
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        def on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")

        canvas.bind_all("<MouseWheel>", on_mousewheel)

        def bind_scroll(event):
            canvas.bind_all("<MouseWheel>", on_mousewheel)

        def unbind_scroll(event):
            canvas.unbind_all("<MouseWheel>")

        canvas.bind("<Enter>", bind_scroll)
        canvas.bind("<Leave>", unbind_scroll)

        for i, row in enumerate(worksheet.iter_rows(min_row=1, min_col=0, values_only=True)):
            for j, value in enumerate(row):
                data = ttk.Label(scrollable_frame, width=20, text=value, anchor='center')
                data.grid(row=i, column=j)
                zp_label = ttk.Label(scrollable_frame, width=20, text="ZP", anchor='center')
                zp_label.grid(row=0, column=0)
                order_label = ttk.Label(scrollable_frame, width=20, text="Costumer Order", anchor='center')
                order_label.grid(row=0, column=1)
                quantity_label = ttk.Label(scrollable_frame, width=20, text="Quantity", anchor='center')
                quantity_label.grid(row=0, column=2)
                done_label = ttk.Label(scrollable_frame, width=20, text="Done", anchor='center')
                done_label.grid(row=0, column=3)

                var = BooleanVar()
                def print_selected(var, val, row_index):
                    if var.get() == 1:
                        root.update_idletasks()
                        # Load the workbook
                        workbook = openpyxl.load_workbook('done.xlsx')

                        # # Create a new sheet
                        # sheet = workbook.create_sheet("Sheet1")

                        # Select the sheet you want to add a record to
                        sheet = workbook['Sheet']

                        # Add a new record to the sheet
                        new_record = [worksheet.cell(row=row_index, column=1).value, worksheet.cell(row=row_index, column=2).value, val]
                        sheet.append(new_record)

                        def delete():
                            # Open the workbook
                            wb = openpyxl.load_workbook('example.xlsx')

                            # Select the worksheet
                            ws = wb['Sheet']

                            # Delete the row
                            ws.delete_rows(row_index)

                            # Save the changes
                            wb.save('example.xlsx')

                        delete()


                        # Save the workbook
                        workbook.save('done.xlsx')

                        # print("Costumer Order:", val)
                        # print("Zlecenie Produkcji:", worksheet.cell(row=row_index+1, column=1).value)
                        # print("Quantity:", worksheet.cell(row=row_index+1, column=2).value)

                        
                done_checkbutton = ttk.Checkbutton(scrollable_frame, variable=var, onvalue=1, offvalue=0, command=lambda var_value=var, val=value, row_index=i: print_selected(var_value, val, row_index))

                done_checkbutton.grid(row=i, column=j+1)
        # Close the workbook
        workbook.close()

        to_do_lf.pack()

        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        separator = ttk.Separator(root, orient='horizontal')
        separator.pack(fill='x', pady=20)

    def done():
        # # Create a new workbook
        # workbook = openpyxl.Workbook()

        # Open the workbook
        workbook = openpyxl.load_workbook('done.xlsx')

        # Get the active worksheet
        worksheet = workbook.active
        done_lf = LabelFrame(root, text='Done', bd=2, relief=RIDGE)
        
        # Create a canvas with a scrollbar
        canvas_done = tkinter.Canvas(done_lf, width=800, height=350)
        scrollbar_done = ttk.Scrollbar(done_lf, orient="vertical", command=canvas_done.yview)
        scrollable_frame_done = ttk.Frame(canvas_done)

        scrollable_frame_done.bind(
            "<Configure>",
            lambda e: canvas_done.configure(
                scrollregion=canvas_done.bbox("all")
            )
        )

        canvas_done.create_window((0, 0), window=scrollable_frame_done, anchor="nw")
        canvas_done.configure(yscrollcommand=scrollbar_done.set)

        def on_mousewheel_done(event):
            canvas_done.yview_scroll(int(-1*(event.delta/120)), "units")

        canvas_done.bind_all("<MouseWheel>", on_mousewheel_done)

        def bind_scroll(event):
            canvas_done.bind_all("<MouseWheel>", on_mousewheel_done)

        def unbind_scroll(event):
            canvas_done.unbind_all("<MouseWheel>")

        canvas_done.bind("<Enter>", bind_scroll)
        canvas_done.bind("<Leave>", unbind_scroll)

        # Read data from the worksheet
        i=0
        for row in worksheet.iter_rows(min_row=0, min_col=0, values_only=True):
            for j in range(0,len(row)):
                e=ttk.Label(scrollable_frame_done,width=20,text=row[j],
                anchor='center')
                e.grid(row=i,column=j)
                zp_label = ttk.Label(scrollable_frame_done, width=20, text="ZP", anchor='center')
                zp_label.grid(row=0, column=0)
                order_label = ttk.Label(scrollable_frame_done, width=20, text="Costumer Order", anchor='center')
                order_label.grid(row=0, column=1)
                quantity_label = ttk.Label(scrollable_frame_done, width=20, text="Quantity", anchor='center')
                quantity_label.grid(row=0, column=2)

            i += 1
            
        # Close the workbook
        workbook.close()

        done_lf.pack()

        scrollbar_done.pack(side="right", fill="y")
        canvas_done.pack(side="left", fill="both", expand=True)
        done_lf.update_idletasks()
        # def refresh():
        #     done_lf.update_idletasks()

    todo()
    done()

    root.mainloop()

main()